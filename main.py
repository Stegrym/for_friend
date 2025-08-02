import os, csv
import openpyxl
from typing import List, Tuple


def get_csv_files(base_dir, base_folder_name) -> List[Tuple[str, str]]:
    """ Ищет .CSV файлы, сохраняет полученные данные в список.

    :param base_dir: Имя директории для поиска.
    :param base_folder_name: Имя текущей папки.
    :return: Cписок с кортежами (имя папки где файл, имя файла)
    """

    csv_files = list()

    # Через рекурсию можно было бы пройти все уровни вложенных папок.
    for item in os.listdir(base_dir):
        try:
            item_path = os.path.join(base_dir, item)
            # Если item файл
            if os.path.isfile(item_path) and item.lower().endswith(".csv"):
                csv_files.append((base_folder_name, item))
            # Если item папка
            elif os.path.isdir(item_path):
                for sub_item in os.listdir(item_path):
                    sub_item_path = os.path.join(item_path, sub_item)
                    if os.path.isfile(sub_item_path) and sub_item.lower().endswith(".csv"):
                        csv_files.append((item, sub_item))
        except PermissionError:
            continue
    return csv_files


def extract_institution(file_path: str) -> str:
    """Извлекает из файла название учреждения между маркерами.

    :param file_path: Путь для открытия файла
    :return: Название или данные об ошибке
    """

    # Маркеры, для выборки
    start_marker = "; - "
    end_marker = "; Email - "

    try:
        # Я не переписываю твой код, но может лучше регулярные выражения?
        # pattern = r";\s*-\s*(.*?);\s*Email\s*-"
        with open(file_path, "r", encoding="UTF-8") as f:

            content = f.read()
            start_idx = content.find(start_marker)

            if start_idx != -1:
                start_idx += len(start_marker)
                end_idx = content.find(end_marker, start_idx)
                if end_idx != -1:
                    return content[start_idx:end_idx]
                else:
                    return "MARKER_ERROR: end_marker not found"
            else:
                return "MARKER_ERROR: start_marker not found"
    except Exception as e:
        return f"ERROR: {str(e)}"


def extract_last_fields(file_path: str, max_result_size: int) -> List:
    """

    :param max_result_size: Ограничение для вывода ответа
    :param file_path: Путь к файлу
    :return: Список с данными колонок
    """
    result = list()
    lines = list()
    START_LINE = 1
    MAX_LINES = 113  # Количество строк

    try:
        with open(file_path, "r", encoding="UTF-8") as f:
            reader = csv.reader(f)
            # Собираем не пустые строки
            for line in reader:
                if any(field.strip() for field in line):
                    lines.append(line)

            lines = lines[START_LINE:START_LINE + MAX_LINES]  # Получили нужное чисто строк

        for line in lines:
            if len(result) >= max_result_size:
                break

            if len(line) > 0:
                last_field = line[-1].strip()

                if last_field:
                    result.append(last_field[-2:])
                elif len(line) > 1:
                    second_last = line[-2].strip()
                    if second_last:
                        result.append(second_last[-2:])
    except Exception:
        pass
    return result


def main():
    base_dir = os.getcwd()
    base_folder_name = os.path.basename(base_dir)
    max_result_size = 101  # Ограничитель ответ

    # Создание Exel таблицы
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Список всех .csv файлов
    csv_files = get_csv_files(base_dir, base_folder_name)

    # РАСПАКОвКА!
    for col_idx, (folder_name, filename) in enumerate(csv_files, 1):

        # Создаём путь
        file_path = os.path.join(
            base_dir,
            folder_name if folder_name != base_folder_name else "",  # Если нужна, то появляется
            filename)

        # Запись данных в Exel строка 1
        ws.cell(row=1, column=col_idx, value=folder_name)

        file_base = os.path.splitext(filename)[0]
        words = file_base.split()
        last_three = " ".join(words[-3:]) if len(words) >= 3 else file_base
        # Запись данных в Exel строка 2
        ws.cell(row=2, column=col_idx, value=last_three)

        institution_text = extract_institution(file_path)
        # Запись данных в Exel строка 3
        ws.cell(row=3, column=col_idx, value=institution_text)

        # Список нужный
        data_rows = extract_last_fields(file_path, max_result_size)

        # Добавление данный в таблицу
        for row_idx, value in enumerate(data_rows[:max_result_size], 4):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Вывод результирующей таблицы
    output_path = os.path.join(base_dir, "result.xlsx")
    wb.save(output_path)
    print(f"Файл успешно создан: {output_path}")


if __name__ == "__main__":
    main()
